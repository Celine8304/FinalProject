import os
import tempfile
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.check_template import CheckTemplate
from app.schemas.template import TemplateResponse
from app.services.template_import_service import import_templates_from_excel

router = APIRouter(prefix="/templates", tags=["模板管理"])


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "").replace(" ", "")


def build_full_field_mapping(excel_path: str):
    """
    从完整字段表中读取：
    控制点 + 控制项 -> 检查方法、推荐值
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = None
    header_map = {}

    required_headers = ["控制点", "控制项", "检查方法", "推荐值", "权重"]

    # 在前 15 行里找表头
    for row_idx in range(1, 16):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        temp_map = {}
        for col_idx, val in enumerate(row_values, start=1):
            if val in required_headers:
                temp_map[val] = col_idx
        if (
    "控制点" in temp_map
    and "控制项" in temp_map
    and "检查方法" in temp_map
    and "推荐值" in temp_map
    and "权重" in temp_map
):
            header_row = row_idx
            header_map = temp_map
            break

    if not header_row:
        raise ValueError("完整字段表未找到表头：控制点 / 控制项 / 检查方法 / 推荐值 / 权重")

    mapping = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        control_point = ws.cell(row=row_idx, column=header_map["控制点"]).value
        control_item = ws.cell(row=row_idx, column=header_map["控制项"]).value
        check_method = ws.cell(row=row_idx, column=header_map["检查方法"]).value
        recommended_value = ws.cell(row=row_idx, column=header_map["推荐值"]).value
        weight = ws.cell(row=row_idx, column=header_map["权重"]).value

        if not control_point or not control_item:
            continue

        key = f"{normalize_text(control_point)}||{normalize_text(control_item)}"
        mapping[key] = {
    "check_method": None if check_method is None else str(check_method).strip(),
    "recommended_value": None if recommended_value is None else str(recommended_value).strip(),
    "weight": None if weight is None else str(weight).strip(),
}

    return mapping


def supplement_template_fields(db, asset_type: str, guide_name: str):
    """
    用完整字段表补充 check_method / recommended_value
    当前先只支持 Redhat Linux
    """
    guide_name_text = str(guide_name).strip()

    if asset_type == "server_storage" and "Redhat" in guide_name_text:
        full_field_filename = "Redhat Linux完整字段.xlsx"
    elif asset_type == "server_storage" and "Ubuntu" in guide_name_text:
        full_field_filename = "Ubuntu完整字段.xlsx"
    elif asset_type == "server_storage" and "银河麒麟" in guide_name_text:
        full_field_filename = "银河麒麟完整字段.xlsx"
    elif asset_type == "server_storage" and "统信" in guide_name_text:
        full_field_filename = "统信完整字段.xlsx"
    elif asset_type == "database" and "人大金仓" in guide_name_text:
        full_field_filename = "人大金仓完整字段.xlsx"
    elif asset_type == "database" and "达梦" in guide_name_text:
        full_field_filename = "达梦完整字段.xlsx"
    else:
        return

    current_dir = Path(__file__).resolve().parent.parent.parent
    excel_path = current_dir / "template_files" / full_field_filename

    if not excel_path.exists():
        print(f"补字段文件不存在：{excel_path}")
        return

    full_mapping = build_full_field_mapping(str(excel_path))

    templates = (
        db.query(CheckTemplate)
        .filter(
            CheckTemplate.asset_type == asset_type,
            CheckTemplate.guide_name == guide_name
        )
        .all()
    )

    update_count = 0

    for item in templates:
        key = f"{normalize_text(item.control_point)}||{normalize_text(item.control_item)}"
        extra = full_mapping.get(key)
        if extra:
            item.check_method = extra["check_method"]
            item.recommended_value = extra["recommended_value"]
            item.weight = extra["weight"]
            update_count += 1

    db.commit()
    print(f"{guide_name} 补字段完成，共更新 {update_count} 条模板记录")


@router.post("/import")
async def import_template_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 文件")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        temp_path = tmp.name

    try:
        result = import_templates_from_excel(temp_path, file.filename, db)

        # 导入缺列模板后，自动用完整字段表补充“检查方法”“推荐值”
        supplement_template_fields(
            db,
            result["asset_type"],
            result["guide_name"]
        )

        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@router.get("/", response_model=list[TemplateResponse])
def list_templates(db: Session = Depends(get_db)):
    templates = db.query(CheckTemplate).order_by(CheckTemplate.id.asc()).all()
    return templates


@router.get("/by-guide/{guide_name}", response_model=list[TemplateResponse])
def list_templates_by_guide_name(guide_name: str, db: Session = Depends(get_db)):
    templates = (
        db.query(CheckTemplate)
        .filter(CheckTemplate.guide_name == guide_name)
        .order_by(CheckTemplate.id.asc())
        .all()
    )
    return templates