from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.asset import Asset
from app.models.check_record import CheckRecord
from app.models.check_template import CheckTemplate
from app.schemas.record import CheckRecordUpdate

from pathlib import Path
import os
from tempfile import NamedTemporaryFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from app.models.asset import Asset

from app.models.check_record import CheckRecord
from app.models.check_template import CheckTemplate


router = APIRouter(prefix="/records", tags=["现场核查记录"])


def compliance_to_score(value):
    if value == "compliant":
        return 5.0
    if value == "partial":
        return 2.5
    if value == "non_compliant":
        return 0.0
    return 0.0


def safe_float(value):
    try:
        return float(value)
    except Exception:
        return 0.0


@router.get("/asset/{asset_id}")
def list_records_by_asset(asset_id: int, db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    records = (
        db.query(CheckRecord, CheckTemplate)
        .join(CheckTemplate, CheckRecord.template_id == CheckTemplate.id)
        .filter(CheckRecord.asset_id == asset_id)
        .order_by(CheckTemplate.id.asc())
        .all()
    )

    result = []
    for record, template in records:
        result.append({
            "id": record.id,
            "project_id": record.project_id,
            "asset_id": record.asset_id,
            "template_id": record.template_id,
            "result_record": record.result_record,
            "compliance_status": record.compliance_status,
            "record_remark": record.record_remark,
            "ai_suggestion": record.ai_suggestion,
            "seq_no": template.seq_no,
            "control_point": template.control_point,
            "control_item": template.control_item,
            "importance_level": template.importance_level,
            "check_content": template.check_content,
            "judgment_standard": template.judgment_standard,
            "template_remark": template.remark,
            "check_method": template.check_method,
            "recommended_value": template.recommended_value,
            "weight": template.weight,
        })

    return result


@router.get("/priority/asset/{asset_id}")
def get_priority_by_asset(asset_id: int, db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    rows = (
        db.query(CheckRecord, CheckTemplate)
        .join(CheckTemplate, CheckRecord.template_id == CheckTemplate.id)
        .filter(CheckRecord.asset_id == asset_id)
        .all()
    )

    result = []
    for record, template in rows:
        if record.compliance_status not in ["partial", "non_compliant"]:
            continue

        score = compliance_to_score(record.compliance_status)
        weight = safe_float(template.weight)
        priority_score = (5 - score) * weight

        result.append({
            "record_id": record.id,
            "seq_no": template.seq_no,
            "control_item": template.control_item,
            "compliance_status": record.compliance_status,
            "weight": weight,
            "score": score,
            "priority_score": round(priority_score, 2),
        })

    def seq_sort_key(item):
        try:
            return int(item["seq_no"])
        except Exception:
            return 999999

    result.sort(key=lambda x: (-x["priority_score"], seq_sort_key(x)))
    for index, item in enumerate(result, start=1):
        item["rank"] = index

    return result

@router.put("/{record_id}")
def update_record(record_id: int, payload: CheckRecordUpdate, db: Session = Depends(get_db)):
    record = db.query(CheckRecord).filter(CheckRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="核查记录不存在")

    if payload.compliance_status is not None:
        allowed_values = ["compliant", "partial", "non_compliant"]
        if payload.compliance_status not in allowed_values:
            raise HTTPException(
                status_code=400,
                detail="符合情况只能是 compliant / partial / non_compliant"
            )
        record.compliance_status = payload.compliance_status

    if payload.result_record is not None:
        record.result_record = payload.result_record

    if payload.record_remark is not None:
        record.record_remark = payload.record_remark

    db.commit()
    db.refresh(record)

    return {
        "message": "核查记录更新成功",
        "record": {
            "id": record.id,
            "project_id": record.project_id,
            "asset_id": record.asset_id,
            "template_id": record.template_id,
            "result_record": record.result_record,
            "compliance_status": record.compliance_status,
            "record_remark": record.record_remark,
            "ai_suggestion": record.ai_suggestion,
        }
    }
def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "").replace(" ", "")


def compliance_to_cn(value):
    mapping = {
        "compliant": "符合",
        "partial": "部分符合",
        "non_compliant": "不符合",
        None: "",
        "": "",
    }
    return mapping.get(value, value)


def get_export_template_path(asset_type: str, guide_name: str) -> Path:
    current_dir = Path(__file__).resolve().parent.parent.parent
    base_dir = current_dir / "template_files"

    if asset_type == "server_storage" and "Redhat" in str(guide_name):
        return base_dir / "RedhatLinux模板.xlsx"
    if asset_type == "database" and "达梦" in str(guide_name):
        return base_dir / "达梦模板.xlsx"

    raise ValueError(f"未找到导出模板：asset_type={asset_type}, guide_name={guide_name}")


@router.get("/export/asset/{asset_id}")
def export_records_by_asset(asset_id: int, db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    template_path = get_export_template_path(asset.asset_type, asset.os_or_db_type)
    if not template_path.exists():
        raise HTTPException(status_code=404, detail=f"导出模板不存在：{template_path.name}")

    wb = load_workbook(template_path)
    ws = wb.active

    # 第7行表头
    header_row = 7
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            header_map[str(val).strip()] = col

    records = (
        db.query(CheckRecord, CheckTemplate)
        .join(CheckTemplate, CheckRecord.template_id == CheckTemplate.id)
        .filter(CheckRecord.asset_id == asset_id)
        .order_by(CheckTemplate.id.asc())
        .all()
    )

    # 三种映射：序号 / 控制点+控制项 / 控制项
    seq_map = {}
    full_map = {}
    item_map = {}

    for record, template in records:
        seq_key = normalize_text(template.seq_no)
        full_key = f"{normalize_text(template.control_point)}||{normalize_text(template.control_item)}"
        item_key = normalize_text(template.control_item)

        if seq_key:
            seq_map[seq_key] = (record, template)
        if full_key:
            full_map[full_key] = (record, template)
        if item_key and item_key not in item_map:
            item_map[item_key] = (record, template)

    filled_count = 0

    # 从第9行开始回填
    start_row = 9
    for row in range(start_row, ws.max_row + 1):
        seq_no = ws.cell(row=row, column=header_map.get("序号", 0)).value if header_map.get("序号") else None
        control_point = ws.cell(row=row, column=header_map.get("控制点", 0)).value if header_map.get("控制点") else None
        control_item = ws.cell(row=row, column=header_map.get("控制项", 0)).value if header_map.get("控制项") else None

        pair = None

        seq_key = normalize_text(seq_no)
        full_key = f"{normalize_text(control_point)}||{normalize_text(control_item)}"
        item_key = normalize_text(control_item)

        if seq_key and seq_key in seq_map:
            pair = seq_map.get(seq_key)
        elif full_key in full_map:
            pair = full_map.get(full_key)
        elif item_key in item_map:
            pair = item_map.get(item_key)

        if not pair:
            continue

        record, template = pair

        if "结果记录" in header_map:
            ws.cell(row=row, column=header_map["结果记录"]).value = record.result_record or ""

        if "符合情况" in header_map:
            ws.cell(row=row, column=header_map["符合情况"]).value = compliance_to_cn(record.compliance_status)

        if "检查方法" in header_map:
            ws.cell(row=row, column=header_map["检查方法"]).value = template.check_method or ""

        if "推荐值" in header_map:
            ws.cell(row=row, column=header_map["推荐值"]).value = template.recommended_value or ""

        filled_count += 1

    print(f"导出回填完成，共回填 {filled_count} 行")

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    wb.save(tmp_path)

    export_name = f"{asset.asset_name}_核查表.xlsx"
    return FileResponse(
        path=tmp_path,
        filename=export_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@router.post("/import/asset/{asset_id}")
async def import_records_from_excel(
    asset_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 文件")

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        temp_path = tmp.name

    try:
        wb = load_workbook(temp_path, data_only=True)
        ws = wb.active

        # 第7行表头
        header_row = 7
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is not None:
                header_map[str(val).strip()] = col

        required_headers = ["结果记录", "符合情况"]
        for h in required_headers:
            if h not in header_map:
                raise HTTPException(status_code=400, detail=f"导入文件缺少列：{h}")

        records = (
            db.query(CheckRecord, CheckTemplate)
            .join(CheckTemplate, CheckRecord.template_id == CheckTemplate.id)
            .filter(CheckRecord.asset_id == asset_id)
            .order_by(CheckTemplate.id.asc())
            .all()
        )

        seq_map, full_map, item_map = build_record_match_maps(records)

        updated_count = 0
        start_row = 9

        for row in range(start_row, ws.max_row + 1):
            seq_no = ws.cell(row=row, column=header_map.get("序号", 0)).value if header_map.get("序号") else None
            control_point = ws.cell(row=row, column=header_map.get("控制点", 0)).value if header_map.get("控制点") else None
            control_item = ws.cell(row=row, column=header_map.get("控制项", 0)).value if header_map.get("控制项") else None

            seq_key = normalize_text(seq_no)
            full_key = f"{normalize_text(control_point)}||{normalize_text(control_item)}"
            item_key = normalize_text(control_item)

            pair = None
            if seq_key and seq_key in seq_map:
                pair = seq_map.get(seq_key)
            elif full_key in full_map:
                pair = full_map.get(full_key)
            elif item_key in item_map:
                pair = item_map.get(item_key)

            if not pair:
                continue

            record, template = pair

            result_record = ws.cell(row=row, column=header_map["结果记录"]).value
            compliance_status_cn = ws.cell(row=row, column=header_map["符合情况"]).value

            record.result_record = None if result_record is None else str(result_record).strip()

            try:
                record.compliance_status = cn_to_compliance(compliance_status_cn)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail=f"第 {row} 行“符合情况”非法，只允许填写：符合 / 部分符合 / 不符合"
                )

            updated_count += 1

        db.commit()

        return {
            "message": "Excel 回填导入成功",
            "asset_id": asset_id,
            "updated_count": updated_count
        }

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

def cn_to_compliance(value):
    mapping = {
        "符合": "compliant",
        "部分符合": "partial",
        "不符合": "non_compliant",
        None: None,
        "": None,
    }
    if value is None:
        return None

    text = str(value).strip()
    if text not in mapping:
        raise ValueError(f"非法的符合情况值：{text}")

    return mapping[text]


def build_record_match_maps(records):
    seq_map = {}
    full_map = {}
    item_map = {}

    for record, template in records:
        seq_key = normalize_text(template.seq_no)
        full_key = f"{normalize_text(template.control_point)}||{normalize_text(template.control_item)}"
        item_key = normalize_text(template.control_item)

        if seq_key:
            seq_map[seq_key] = (record, template)
        if full_key:
            full_map[full_key] = (record, template)
        if item_key and item_key not in item_map:
            item_map[item_key] = (record, template)

    return seq_map, full_map, item_map

