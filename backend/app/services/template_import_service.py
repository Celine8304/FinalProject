from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.check_template import CheckTemplate


HEADER_ROW = 7
DATA_START_ROW = 9


def get_file_config(filename: str) -> dict:
    """
    根据文件名确定资产类型和作业指导书类型
    第一版只支持两个模板文件：
    - RedhatLinux模板.xlsx
    - 达梦模板.xlsx
    """
    lower_name = filename.lower()

    if "redhat" in lower_name:
        return {
            "asset_type": "server_storage",
            "guide_name": "Redhat Linux",
        }

    if "达梦" in filename or "dm" in lower_name:
        return {
            "asset_type": "database",
            "guide_name": "达梦",
        }

    raise ValueError("暂时只支持 RedhatLinux模板.xlsx 和 达梦模板.xlsx")


def build_column_mapping(header_values: list[str]) -> dict:
    """
    根据第7行表头，找到各列下标
    """
    mapping = {}

    for idx, header in enumerate(header_values):
        text = str(header).strip() if header is not None else ""

        if "序号" in text:
            mapping["seq_no"] = idx
        elif "控制点" in text:
            mapping["control_point"] = idx
        elif "控制项" in text:
            mapping["control_item"] = idx
        elif "备注" in text:
            mapping["remark"] = idx
        elif "重要程度" in text:
            mapping["importance_level"] = idx
        elif "检查内容" in text:
            mapping["check_content"] = idx
        elif "判断标准" in text:
            mapping["judgment_standard"] = idx

    return mapping


def safe_get(row_values: list, mapping: dict, field_name: str):
    idx = mapping.get(field_name)
    if idx is None or idx >= len(row_values):
        return None

    value = row_values[idx]
    if value is None:
        return None

    return str(value).strip()


def import_templates_from_excel(file_path: str, original_filename: str, db: Session) -> dict:
    config = get_file_config(original_filename)

    asset_type = config["asset_type"]
    guide_name = config["guide_name"]

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # 这两个模板文件第一版默认只读活动sheet

    header_values = [cell.value for cell in ws[HEADER_ROW]]
    column_mapping = build_column_mapping(header_values)

    required_fields = [
        "seq_no",
        "control_point",
        "control_item",
        "check_content",
        "judgment_standard",
    ]
    missing_fields = [field for field in required_fields if field not in column_mapping]
    if missing_fields:
        raise ValueError(f"模板表头缺少关键列: {missing_fields}")

    # 为避免重复导入，先删除同 guide_name 的旧模板
    db.query(CheckTemplate).filter(CheckTemplate.guide_name == guide_name).delete()

    imported_count = 0

    # 从第9行开始读，第8行直接跳过
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_idx]]

        seq_no = safe_get(row_values, column_mapping, "seq_no")
        control_point = safe_get(row_values, column_mapping, "control_point")
        control_item = safe_get(row_values, column_mapping, "control_item")
        remark = safe_get(row_values, column_mapping, "remark")
        importance_level = safe_get(row_values, column_mapping, "importance_level")
        check_content = safe_get(row_values, column_mapping, "check_content")
        judgment_standard = safe_get(row_values, column_mapping, "judgment_standard")

        # 空白行跳过
        if not any([seq_no, control_point, control_item, check_content, judgment_standard]):
            continue

        template = CheckTemplate(
            asset_type=asset_type,
            guide_name=guide_name,
            sheet_name=ws.title,
            seq_no=seq_no,
            control_point=control_point,
            control_item=control_item,
            importance_level=importance_level,
            check_content=check_content,
            judgment_standard=judgment_standard,
            remark=remark,
        )
        db.add(template)
        imported_count += 1

    db.commit()

    return {
        "message": "模板导入成功",
        "asset_type": asset_type,
        "guide_name": guide_name,
        "sheet_name": ws.title,
        "imported_count": imported_count,
        "header_row": HEADER_ROW,
        "data_start_row": DATA_START_ROW,
    }